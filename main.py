import os
import json
import uuid
from datetime import datetime, timedelta
from io import BytesIO
from typing import Any, Dict, List, Tuple, Optional

import pandas as pd
from fastapi import FastAPI, HTTPException, Query
from fastapi.responses import JSONResponse
from google.oauth2 import service_account
from googleapiclient.discovery import build
from googleapiclient.http import MediaIoBaseDownload
from zoneinfo import ZoneInfo

from openpyxl import load_workbook

import smtplib
from email.message import EmailMessage

# 要約（別ファイル）
from summary import generate_summary

# ----------------------------
# FastAPI App
# ----------------------------
app = FastAPI()

# ----------------------------
# 簡易ロック＆簡易台帳（練習用：プロセス内メモリ）
# ※本番はDrive/DBなど永続化へ
# ----------------------------
RUN_LOCK: Dict[str, datetime] = {}           # as_of_date -> lock acquired time
RUN_SUCCESS: Dict[str, Dict[str, Any]] = {}  # as_of_date -> success record

LOCK_TTL_SECONDS = 60 * 30  # 30分（保険）

# ----------------------------
# 設定
# ----------------------------
REQUIRED_FILES = [
    "CPH.xlsx",
    "AHT.xlsx",
    "ATT.xlsx",
    "ACW.xlsx",
    "CPD.xlsx",
    "着座比率.xlsx",
    "稼働率.xlsx",
]

METRIC_BY_FILENAME = {
    "CPH.xlsx": "CPH",
    "AHT.xlsx": "AHT",
    "ATT.xlsx": "ATT",
    "ACW.xlsx": "ACW",
    "CPD.xlsx": "CPD",
    "着座比率.xlsx": "着座比率",
    "稼働率.xlsx": "稼働率",
}

# INPUT共通列（実データ前提）
BASE_COLS = ["日付", "agent_id", "氏名", "勤務区分", "実働時間(h)", "CPD目標"]

# テンプレシート名（違うならここだけ変える）
SHEET_FACT_DAILY = "Fact_Daily"
SHEET_FACT_LONG = "Fact_Long"


# ----------------------------
# API
# ----------------------------
@app.get("/health")
def health():
    return {"status": "ok", "version": "2026-02-21-summary-mailbody-01"}


# ----------------------------
# Drive（読み取り専用）
# ----------------------------
def _get_drive_service():
    sa_json = os.environ.get("GCP_SA_JSON")
    if not sa_json:
        raise RuntimeError("Missing env: GCP_SA_JSON")

    sa_info = json.loads(sa_json)
    credentials = service_account.Credentials.from_service_account_info(
        sa_info,
        scopes=["https://www.googleapis.com/auth/drive.readonly"],
    )
    return build("drive", "v3", credentials=credentials, cache_discovery=False)


def _find_child_folder_by_name(drive, parent_folder_id: str, folder_name: str) -> Optional[Dict[str, str]]:
    q = (
        f"'{parent_folder_id}' in parents and "
        f"mimeType = 'application/vnd.google-apps.folder' and "
        f"name = '{folder_name}' and trashed = false"
    )
    res = drive.files().list(q=q, fields="files(id,name)", pageSize=10).execute()
    files = res.get("files", [])
    return files[0] if files else None


def _list_child_files(drive, parent_folder_id: str, page_size: int = 200) -> List[Dict[str, Any]]:
    q = f"'{parent_folder_id}' in parents and trashed = false"
    res = drive.files().list(
        q=q,
        fields="files(id,name,mimeType,modifiedTime)",
        pageSize=page_size,
        orderBy="name",
    ).execute()
    return res.get("files", [])


def _download_file_bytes(drive, file_id: str) -> bytes:
    fh = BytesIO()
    request = drive.files().get_media(fileId=file_id)
    downloader = MediaIoBaseDownload(fh, request)
    done = False
    while not done:
        _, done = downloader.next_chunk()
    return fh.getvalue()


# ----------------------------
# Excel -> DataFrame（Phase2）
# ----------------------------
def _read_excel_from_bytes(xbytes: bytes, sheet_name: str = "Data") -> pd.DataFrame:
    bio = BytesIO(xbytes)
    try:
        return pd.read_excel(bio, sheet_name=sheet_name)
    except ValueError:
        bio.seek(0)
        return pd.read_excel(bio)  # fallback


def _clean_columns(df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy()
    df.columns = [str(c).strip().replace("　", " ") for c in df.columns]
    return df


def _normalize_common_columns(df: pd.DataFrame) -> pd.DataFrame:
    df = _clean_columns(df)

    rename_map = {}
    for c in df.columns:
        c2 = str(c).strip()
        if c2 in ["エージェントID", "エージェントId", "agent_id", "AgentID", "AGENT_ID", "ID"]:
            rename_map[c] = "agent_id"

    df = df.rename(columns=rename_map)

    missing = [c for c in BASE_COLS if c not in df.columns]
    if missing:
        raise ValueError(f"Missing base columns: {missing}. found={list(df.columns)}")

    df["agent_id"] = df["agent_id"].astype(str).str.strip()
    return df


def _to_numeric_series(s: pd.Series) -> pd.Series:
    x = s.astype(str)
    x = x.str.replace("%", "", regex=False).str.replace(",", "", regex=False).str.strip()
    return pd.to_numeric(x, errors="coerce")


def _extract_metric_series(raw_df: pd.DataFrame, metric_name: str) -> pd.DataFrame:
    df = _normalize_common_columns(raw_df)
    candidates = [c for c in df.columns if c not in BASE_COLS]

    if metric_name in df.columns:
        metric_col = metric_name
    else:
        if len(candidates) != 1:
            raise ValueError(
                f"Cannot uniquely detect metric column for metric={metric_name}. "
                f"candidates={candidates}"
            )
        metric_col = candidates[0]

    out = df[["日付", "agent_id", metric_col]].copy()
    out = out.rename(columns={metric_col: metric_name})
    out[metric_name] = _to_numeric_series(out[metric_name])

    dup = out.duplicated(subset=["日付", "agent_id"], keep=False)
    if dup.any():
        sample = out.loc[dup, ["日付", "agent_id"]].head(5).to_dict(orient="records")
        raise ValueError(f"Duplicate keys in metric={metric_name} on (日付,agent_id). sample={sample}")

    return out


def _build_fact_daily_and_long(
    raw_by_metric: Dict[str, pd.DataFrame],
    as_of_date: str,
) -> Tuple[pd.DataFrame, pd.DataFrame]:
    if "CPD" not in raw_by_metric:
        raise ValueError("CPD.xlsx is required as base but not provided.")

    base = _normalize_common_columns(raw_by_metric["CPD"])
    fact_daily = base[BASE_COLS].copy()

    metric_names = list(raw_by_metric.keys())

    for metric in metric_names:
        metric_df = _extract_metric_series(raw_by_metric[metric], metric)
        fact_daily = fact_daily.merge(
            metric_df,
            on=["日付", "agent_id"],
            how="left",
            validate="one_to_one",
        )

    fact_long = fact_daily.melt(
        id_vars=BASE_COLS,
        value_vars=metric_names,
        var_name="metric",
        value_name="actual",
    )
    fact_long["as_of_date"] = as_of_date
    fact_long["work_flag"] = (fact_long["勤務区分"].astype(str).str.strip() != "休み").astype(int)

    return fact_daily, fact_long


# ----------------------------
# openpyxl（テンプレに書き込み）
# ----------------------------
def _clear_worksheet(ws):
    mr = ws.max_row
    mc = ws.max_column
    if mr < 1 or mc < 1:
        return
    for r in range(1, mr + 1):
        for c in range(1, mc + 1):
            ws.cell(row=r, column=c).value = None


def _write_df_to_sheet(ws, df: pd.DataFrame):
    cols = list(df.columns)
    for c, name in enumerate(cols, start=1):
        ws.cell(row=1, column=c).value = name

    values = df.where(pd.notnull(df), None).values.tolist()
    for r_idx, row in enumerate(values, start=2):
        for c_idx, v in enumerate(row, start=1):
            ws.cell(row=r_idx, column=c_idx).value = v


def _build_output_excel_bytes(template_bytes: bytes, fact_daily: pd.DataFrame, fact_long: pd.DataFrame) -> bytes:
    wb = load_workbook(BytesIO(template_bytes))

    if SHEET_FACT_DAILY not in wb.sheetnames:
        raise ValueError(f"Template missing sheet: {SHEET_FACT_DAILY}. sheets={wb.sheetnames}")
    if SHEET_FACT_LONG not in wb.sheetnames:
        raise ValueError(f"Template missing sheet: {SHEET_FACT_LONG}. sheets={wb.sheetnames}")

    ws_daily = wb[SHEET_FACT_DAILY]
    ws_long = wb[SHEET_FACT_LONG]

    _clear_worksheet(ws_daily)
    _write_df_to_sheet(ws_daily, fact_daily)

    _clear_worksheet(ws_long)
    _write_df_to_sheet(ws_long, fact_long)

    out = BytesIO()
    wb.save(out)
    return out.getvalue()


# ----------------------------
# Mail（Mailtrap含む SMTP 添付送信）
# ----------------------------
def _send_mail_with_attachment(
    subject: str,
    body: str,
    attachment_bytes: bytes,
    attachment_filename: str,
):
    smtp_host = os.environ.get("SMTP_HOST")
    smtp_port = int(os.environ.get("SMTP_PORT", "587"))
    smtp_user = os.environ.get("SMTP_USER")
    smtp_pass = os.environ.get("SMTP_PASS")

    mail_to = os.environ.get("MAIL_TO")
    mail_from = os.environ.get("MAIL_FROM") or smtp_user

    missing = [k for k, v in {
        "SMTP_HOST": smtp_host,
        "SMTP_USER": smtp_user,
        "SMTP_PASS": smtp_pass,
        "MAIL_TO": mail_to,
    }.items() if not v]
    if missing:
        raise RuntimeError(f"Missing env for mail: {missing}")

    to_list = [x.strip() for x in mail_to.split(",") if x.strip()]
    if not to_list:
        raise RuntimeError("MAIL_TO is empty")

    msg = EmailMessage()
    msg["Subject"] = subject
    msg["From"] = mail_from
    msg["To"] = ", ".join(to_list)
    msg.set_content(body)

    msg.add_attachment(
        attachment_bytes,
        maintype="application",
        subtype="vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        filename=attachment_filename,
    )

    with smtplib.SMTP(smtp_host, smtp_port, timeout=30) as s:
        s.ehlo()
        s.starttls()
        s.ehlo()
        s.login(smtp_user, smtp_pass)
        s.send_message(msg)


# ----------------------------
# Main Endpoint
# ----------------------------
@app.post("/run_daily_close")
def run_daily_close(
    target_date: Optional[str] = Query(default=None, description="YYYY-MM-DD. default = yesterday(JST)")
):
    input_folder_id = os.environ.get("DRIVE_INPUT_FOLDER_ID")
    template_file_id = os.environ.get("DRIVE_TEMPLATE_FILE_ID")

    if not input_folder_id:
        raise HTTPException(status_code=500, detail="Missing env: DRIVE_INPUT_FOLDER_ID")
    if not template_file_id:
        raise HTTPException(status_code=500, detail="Missing env: DRIVE_TEMPLATE_FILE_ID")

    jst = ZoneInfo("Asia/Tokyo")
    if target_date:
        as_of_date = target_date.strip()
    else:
        as_of_date_dt = datetime.now(jst).date() - timedelta(days=1)
        as_of_date = as_of_date_dt.strftime("%Y-%m-%d")

    # 既に成功済み（練習用：メモリ台帳）
    if as_of_date in RUN_SUCCESS:
        prev = RUN_SUCCESS[as_of_date]
        return {
            "result": "already_sent",
            "as_of_date": as_of_date,
            "run_id": prev.get("run_id"),
            "finished_at": prev.get("finished_at"),
            "message": "Already sent for this date (memory ledger).",
        }

    # ロックTTL掃除
    now = datetime.now(jst)
    expired_keys = []
    for k, t in RUN_LOCK.items():
        if (now - t).total_seconds() > LOCK_TTL_SECONDS:
            expired_keys.append(k)
    for k in expired_keys:
        RUN_LOCK.pop(k, None)

    # 実行中ロック
    if as_of_date in RUN_LOCK:
        return {
            "result": "running",
            "as_of_date": as_of_date,
            "message": "Job is already running (memory lock).",
        }

    # ロック取得
    RUN_LOCK[as_of_date] = now
    run_id = str(uuid.uuid4())

    try:
        drive = _get_drive_service()

        # Phase1: 前日フォルダ取得
        daily_folder = _find_child_folder_by_name(drive, input_folder_id, as_of_date)
        if not daily_folder:
            return {
                "result": "input_not_ready",
                "phase": "phase1_find_folder",
                "as_of_date": as_of_date,
                "run_id": run_id,
                "detail": f"Daily folder not found: {as_of_date}",
            }

        daily_folder_id = daily_folder["id"]
        children = _list_child_files(drive, daily_folder_id, page_size=200)

        found_file_names = sorted(
            [
                f["name"]
                for f in children
                if f.get("mimeType") != "application/vnd.google-apps.folder"
            ]
        )

        found_set = set(found_file_names)
        missing_files = [name for name in REQUIRED_FILES if name not in found_set]

        if missing_files:
            return {
                "result": "input_not_ready",
                "phase": "phase1_validate_inputs",
                "as_of_date": as_of_date,
                "run_id": run_id,
                "missing_files": missing_files,
                "found_files": found_file_names,
            }

        # Phase2: Data読み込み
        file_id_by_name = {
            f["name"]: f["id"]
            for f in children
            if f.get("mimeType") != "application/vnd.google-apps.folder"
        }

        raw_by_metric: Dict[str, pd.DataFrame] = {}
        for filename in REQUIRED_FILES:
            metric = METRIC_BY_FILENAME[filename]
            file_id = file_id_by_name[filename]
            xbytes = _download_file_bytes(drive, file_id)
            df = _read_excel_from_bytes(xbytes, sheet_name="Data")
            raw_by_metric[metric] = df

        fact_daily, fact_long = _build_fact_daily_and_long(raw_by_metric, as_of_date)

        # Phase3: テンプレ取得 -> Excel生成
        template_bytes = _download_file_bytes(drive, template_file_id)
        output_excel_bytes = _build_output_excel_bytes(template_bytes, fact_daily, fact_long)

        # 要約生成（落ちても送信は止めない）
        try:
            summary = generate_summary(fact_daily, as_of_date)
        except Exception as e:
            summary = f"要約生成に失敗しました: {type(e).__name__}: {e}"

        # メール送信（添付）
        attach_name = f"{as_of_date}_前日確定版_実績.xlsx"
        subject = f"[前日確定版] {as_of_date} 実績レポート"

        body = (
            f"{as_of_date} の前日確定版レポートを生成しました。\n"
            f"添付ファイルをご確認ください。\n\n"
            f"▼ 5行要約\n"
            f"{summary}\n"
        )

        _send_mail_with_attachment(
            subject=subject,
            body=body,
            attachment_bytes=output_excel_bytes,
            attachment_filename=attach_name,
        )

        # 成功を簡易台帳に保存（練習用）
        RUN_SUCCESS[as_of_date] = {
            "run_id": run_id,
            "finished_at": datetime.now(jst).isoformat(),
            "fact_daily_rows": int(len(fact_daily)),
            "fact_long_rows": int(len(fact_long)),
            "attachment_filename": attach_name,
        }

        return {
            "result": "success",
            "phase": "phase3_mail_sent",
            "as_of_date": as_of_date,
            "run_id": run_id,
            "fact_daily_rows": int(len(fact_daily)),
            "fact_long_rows": int(len(fact_long)),
            "attachment_filename": attach_name,
        }

    except Exception as e:
        return JSONResponse(
            status_code=200,
            content={
                "result": "failed",
                "as_of_date": as_of_date,
                "run_id": run_id,
                "error_type": type(e).__name__,
                "error_message": str(e),
            },
        )

    finally:
        RUN_LOCK.pop(as_of_date, None)
